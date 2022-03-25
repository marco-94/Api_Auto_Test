from openpyxl import load_workbook
from config.config import EXCEL_PATH
import json


def execlstr(row, column):
    """读取string，返回string"""
    wb = load_workbook(filename=EXCEL_PATH)
    names = wb.sheetnames
    sheet = wb[names[0]]
    execldata = sheet.cell(row=row, column=column).value

    return execldata


def execljson(row, column):
    """读取string，返回json"""
    wb = load_workbook(filename=EXCEL_PATH)
    names = wb.sheetnames
    sheet = wb[names[0]]
    execl_json = json.loads(sheet.cell(row=row, column=column).value)

    return execl_json


def getRowsColumns():
    """获取用例总数，单用例总接口数"""
    wb = load_workbook(filename=EXCEL_PATH)
    names = wb.sheetnames
    sheet = wb[names[0]]
    CaseTotal = sheet.max_row - 4
    interfacesTotal = int((sheet.max_column - 3) / 6)
    return CaseTotal, interfacesTotal


# def getRowsColumns():
#     """获取用例总数，单用例总接口数"""
#     wb = load_workbook(filename=EXCEL_PATH)
#     names = wb.sheetnames
#     sheet = wb[names[0]]
#     RowsTotal = sheet.max_row
#     interfacesTotal = int((sheet.max_column - 3) / 6)
#     return RowsTotal, interfacesTotal
